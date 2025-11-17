"""
Export Utilities for PDF and Excel Reports
Provides functions to export data to PDF and Excel formats
"""
import io
from datetime import datetime
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def render_to_pdf(template_src, context_dict={}):
    """
    Render a Django template to PDF
    
    Args:
        template_src: Template file path
        context_dict: Context dictionary for template
    
    Returns:
        HttpResponse with PDF content or None on error
    """
    template = get_template(template_src)
    html = template.render(context_dict)
    result = io.BytesIO()
    
    # Convert HTML to PDF
    pdf = pisa.pisaDocument(io.BytesIO(html.encode("UTF-8")), result)
    
    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return None


def export_to_excel(data, headers, filename, sheet_name="Sheet1", title=None):
    """
    Export data to Excel format with styling
    
    Args:
        data: List of lists/tuples containing row data
        headers: List of column headers
        filename: Name of the file to download
        sheet_name: Name of the worksheet
        title: Optional title for the sheet
    
    Returns:
        HttpResponse with Excel content
    """
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    title_font = Font(bold=True, size=16, color="1F4E78")
    title_alignment = Alignment(horizontal="center", vertical="center")
    
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    current_row = 1
    
    # Add title if provided
    if title:
        ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
        title_cell = ws['A1']
        title_cell.value = title
        title_cell.font = title_font
        title_cell.alignment = title_alignment
        current_row = 2
        
        # Add generation date
        ws.merge_cells(f'A2:{get_column_letter(len(headers))}2')
        date_cell = ws['A2']
        date_cell.value = f"Generated on: {datetime.now().strftime('%B %d, %Y %I:%M %p')}"
        date_cell.alignment = Alignment(horizontal="center")
        current_row = 4
    
    # Add headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border_style
    
    # Add data
    for row_num, row_data in enumerate(data, current_row + 1):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.border = border_style
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        
        for cell in ws[column_letter]:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Cap at 50
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


# Specific export functions for different reports

def export_inventory_to_excel(inventory_items, owner_name):
    """Export inventory list to Excel"""
    headers = ['Product Name', 'Supplier', 'Quantity', 'Unit', 'Low Stock Threshold', 'Status', 'Last Restocked']
    
    data = []
    for item in inventory_items:
        status = 'Out of Stock' if item.quantity == 0 else ('Low Stock' if item.is_low_stock else 'In Stock')
        last_restocked = item.last_restocked.strftime('%Y-%m-%d %H:%M') if item.last_restocked else 'Never'
        
        data.append([
            item.product.name,
            item.product.supplier.name if item.product.supplier else 'N/A',
            item.quantity,
            item.product.unit,
            item.low_stock_threshold,
            status,
            last_restocked
        ])
    
    filename = f"inventory_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    title = f"Inventory Report - {owner_name}"
    
    return export_to_excel(data, headers, filename, "Inventory", title)


def export_stock_out_to_excel(stock_outs, owner_name):
    """Export stock out report to Excel"""
    headers = ['Date', 'Product', 'Quantity', 'Unit', 'Reason', 'Payment Method', 'Payment Status', 'Processed By', 'Remarks']
    
    data = []
    for stock_out in stock_outs:
        data.append([
            stock_out.created_at.strftime('%Y-%m-%d %H:%M'),
            stock_out.product.name,
            stock_out.quantity,
            stock_out.product.unit,
            stock_out.get_reason_display(),
            stock_out.get_payment_method_display(),
            stock_out.get_payment_status_display(),
            stock_out.processed_by.first_name if stock_out.processed_by else 'N/A',
            stock_out.remarks or ''
        ])
    
    filename = f"stock_out_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    title = f"Stock Out Report - {owner_name}"
    
    return export_to_excel(data, headers, filename, "Stock Out", title)


def export_orders_to_excel(orders, user_name, user_type="Owner"):
    """Export orders to Excel"""
    headers = ['Order ID', 'Date', 'Product', 'Quantity', 'Unit Price', 'Total Amount', 
               'Order Status', 'Delivery Status', 'Payment Method', 'Payment Status']
    
    data = []
    for order in orders:
        data.append([
            f"#{order.id}",
            order.created_at.strftime('%Y-%m-%d %H:%M'),
            order.supplier_product.name,
            order.quantity,
            f"₱{order.unit_price}",
            f"₱{order.total_amount}",
            order.get_order_status_display(),
            order.get_delivery_status_display(),
            order.get_payment_method_display(),
            order.get_payment_status_display()
        ])
    
    filename = f"orders_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    title = f"Orders Report - {user_name} ({user_type})"
    
    return export_to_excel(data, headers, filename, "Orders", title)


def export_supplier_products_to_excel(products, supplier_name):
    """Export supplier products to Excel"""
    headers = ['Product Name', 'Category', 'Unit Price', 'Available Stock', 'Unit', 
               'Minimum Order', 'Low Stock Threshold', 'Status']
    
    data = []
    for product in products:
        status = 'Out of Stock' if product.is_out_of_stock else ('Low Stock' if product.is_low_stock else 'In Stock')
        
        data.append([
            product.name,
            product.get_category_display(),
            f"₱{product.unit_price}",
            product.available_stock,
            product.unit,
            product.minimum_order_quantity,
            product.low_stock_threshold,
            status
        ])
    
    filename = f"supplier_products_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    title = f"Product Inventory - {supplier_name}"
    
    return export_to_excel(data, headers, filename, "Products", title)


def export_payment_verification_to_excel(orders, user_name):
    """Export payment verification report to Excel"""
    headers = ['Order ID', 'Date', 'Product', 'Amount', 'Payment Method', 
               'Payment Status', 'Verified', 'Verified By', 'Verified Date']
    
    data = []
    for order in orders:
        verified_date = order.verified_date.strftime('%Y-%m-%d %H:%M') if order.verified_date else 'Not Verified'
        verified_by = order.verified_by.first_name if order.verified_by else 'N/A'
        
        data.append([
            f"#{order.id}",
            order.created_at.strftime('%Y-%m-%d'),
            order.supplier_product.name,
            f"₱{order.total_amount}",
            order.get_payment_method_display(),
            order.get_payment_status_display(),
            'Yes' if order.payment_verified else 'No',
            verified_by,
            verified_date
        ])
    
    filename = f"payment_verification_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    title = f"Payment Verification Report - {user_name}"
    
    return export_to_excel(data, headers, filename, "Payment Verification", title)
